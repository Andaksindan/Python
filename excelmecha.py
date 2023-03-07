import openpyxl

# Open the workbook
workbook = openpyxl.load_workbook('MakkalSelvan.xlsx')

# Get the active worksheet
worksheet = workbook.active

# Accessing Cells
cell = worksheet['A3']
print(cell.value)

# Accessing Rows and Columns
row = worksheet[1]
print(row)

column = worksheet['A']
print(column)

# Iterating through Rows
for row in worksheet.iter_rows():
    print(row)

# Iterating through Columns
for column in worksheet.iter_cols():
    print(column)

# Writing to Cells
worksheet['A1'] = 'People man'

# Saving Changes to the Workbook
workbook.save('MakkalSelvan.xlsx')
